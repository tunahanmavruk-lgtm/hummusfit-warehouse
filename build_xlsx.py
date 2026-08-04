import json
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

lane_plan = json.load(open('/home/claude/hummusfit-warehouse/lane_plan.json'))

wb = Workbook()

ARIAL = 'Arial'
INK = '111417'
TEAL = '2BBFAA'
ORANGE = 'E8612C'
GRAY = '767c85'
LIGHT = 'F2F2F2'

def style_header(ws, row, ncols, fill=INK, font_color='FFFFFF'):
    for c in range(1, ncols+1):
        cell = ws.cell(row=row, column=c)
        cell.font = Font(name=ARIAL, bold=True, color=font_color, size=11)
        cell.fill = PatternFill('solid', fgColor=fill)
        cell.alignment = Alignment(vertical='center', wrap_text=True)

thin = Side(style='thin', color='DDDDDD')
border = Border(left=thin, right=thin, top=thin, bottom=thin)

# ---------------- Summary sheet ----------------
ws = wb.active
ws.title = 'Read Me First'
ws.column_dimensions['A'].width = 100
rows = [
    ('Hummus Fit — Walk-In Picking Floor Plan: Crate-by-Crate Build Sheet', True, 15),
    ('', False, 11),
    ('ROOM', True, 12),
    ('Walk-in dimensions: 60\'-11 5/8" x 21\'-11 1/2", 10\' ceiling. Blue roll-up door is the single in/out point for the pick cart.', False, 11),
    ('Layout: 6 rows total in 2 sections — K1, K2, K3 (Bakery & Snacks) and M1, M2, M3 (Meals). 27 lane positions per row, 162 total floor positions.', False, 11),
    ('Each row is 3 double-stacked crate blocks with a walking aisle between — the corrected 6-row layout, not the earlier flawed 8-row version.', False, 11),
    ('', False, 11),
    ('CRATE', True, 12),
    ('Storage crate: 21.65"L x 15.70"W x 10.70"H, self-stacking lid grooves. Max stack height per lane: 5 crates (fits the 10\' ceiling).', False, 11),
    ('Active pick face: bottom 4 crates in a stack. The 5th (top) crate is reserve only — restock crew brings it down, pickers never reach for it.', False, 11),
    ('', False, 11),
    ('DEMAND DATA — READ THIS BEFORE TRUSTING THE NUMBERS', True, 12),
    ('Crate counts below use REAL fulfillment data wherever available: a parsed 180-order Monday shipout batch (178 orders with usable line items),', False, 11),
    ('cross-referenced against 30-day blended Shopify sales (all 16 locations/channels, not warehouse-specific).', False, 11),
    ('Finding: blended Shopify sales overstate real per-day warehouse pick volume by roughly 3-6x for the large majority of meal SKUs (100 of 120', False, 11),
    ('products diverged by more than 2.5x). This is why earlier capacity estimates showed a shortfall (~183 positions needed vs ~162 available) —', False, 11),
    ('that estimate used blended sales as the demand proxy. Rebuilt on real order data, everything fits comfortably: 44/81 Bakery positions used,', False, 11),
    ('79/81 Meals positions used, zero overflow.', False, 11),
    ('CAVEAT: the real-order sample is ONE Monday. Day-of-week variance is not accounted for. 17 slower-moving products had zero matches in that', False, 11),
    ('single day and fall back to the blended estimate (flagged in the "Demand Source" column on each section tab) — treat those numbers as weaker.', False, 11),
    ('Recommend collecting a few more shipout batches (a weekday + a weekend) before treating this as final.', False, 11),
    ('', False, 11),
    ('SLOTTING LOGIC', True, 12),
    ('1. Golden zone: within each section, products are ranked by daily demand (real data preferred) and placed fastest-first — Row 1 positions', False, 11),
    ('   nearest that section\'s aisle entrance, working back through Row 2 and Row 3 for slower movers.', False, 11),
    ('2. Order-affinity clustering: within each row, products frequently ordered together (from the real order batch) are placed in adjacent lanes,', False, 11),
    ('   so a picker filling one order walks less distance — this re-orders position WITHIN a row without breaking the row\'s overall velocity tier.', False, 11),
    ('3. Multi-lane SKUs: if a product needs more crates than one lane can hold (5 crates), it gets a second adjacent lane. Uncommon under real', False, 11),
    ('   demand — most SKUs now fit in a single lane.', False, 11),
    ('', False, 11),
    ('HOW TO USE THE OTHER TABS', True, 12),
    ('"Bakery & Snacks" and "Meals" tabs are the actual build instructions — one row per floor lane, in walk order (K1-01 first, K1-02 next, etc).', False, 11),
    ('Tape a location code on the floor/shelf, put that product\'s crates there, stack up to the crate count shown (5-crate max per lane).', False, 11),
]
r = 1
for text, bold, size in rows:
    cell = ws.cell(row=r, column=1, value=text)
    cell.font = Font(name=ARIAL, bold=bold, size=size, color=INK if not bold else TEAL if size==15 else INK)
    cell.alignment = Alignment(wrap_text=True, vertical='top')
    ws.row_dimensions[r].height = 34 if size == 15 else (18 if bold else 16)
    r += 1

# ---------------- Section tabs ----------------
HEADERS = ['Lane Code', 'Row', 'Position in Row', 'Product', 'Category', 'Crates to Stock (1-day cover)', 'Daily Demand (units)', 'Demand Source', 'Data Confidence Flag']

def cat_label(c):
    return {'muffin': 'Muffin', 'oats': 'Oats', 'snack': 'Snack', 'meal': 'Meal'}.get(c, c)

def write_section(sheet_name, lanes):
    ws = wb.create_sheet(sheet_name)
    for i, h in enumerate(HEADERS, start=1):
        ws.cell(row=1, column=i, value=h)
    style_header(ws, 1, len(HEADERS))
    ws.freeze_panes = 'A2'

    widths = [12, 8, 14, 40, 12, 22, 18, 32, 20]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    for r_idx, lane in enumerate(lanes, start=2):
        row_label, pos = lane['code'].split('-')
        vals = [
            lane['code'], row_label, int(pos), lane['product'], cat_label(lane['category']),
            lane['crates_needed'], lane['daily_avg'], lane['demand_source'],
            'CHECK — real vs blended diverge >2.5x, single-day sample' if lane['divergence_flag'] else ('Fallback — no real-order match' if 'blended' in lane['demand_source'] else 'Good — real order data')
        ]
        for c_idx, v in enumerate(vals, start=1):
            cell = ws.cell(row=r_idx, column=c_idx, value=v)
            cell.font = Font(name=ARIAL, size=10.5, color=INK)
            cell.border = border
            if c_idx in (6, 7):
                cell.alignment = Alignment(horizontal='center')
            if r_idx % 2 == 0:
                cell.fill = PatternFill('solid', fgColor=LIGHT)
        # flag divergent/fallback rows visibly
        flag_cell = ws.cell(row=r_idx, column=9)
        if lane['divergence_flag']:
            flag_cell.font = Font(name=ARIAL, size=10.5, color=ORANGE, bold=True)
        elif 'blended' in lane['demand_source']:
            flag_cell.font = Font(name=ARIAL, size=10.5, color=GRAY, italic=True)
        else:
            flag_cell.font = Font(name=ARIAL, size=10.5, color=TEAL, bold=True)

    ws.row_dimensions[1].height = 30
    return ws

write_section('Bakery & Snacks (K1-K3)', lane_plan['bakery']['lanes'])
write_section('Meals (M1-M3)', lane_plan['meals']['lanes'])

out_path = '/home/claude/hummusfit-warehouse/Hummus_Fit_Crate_Build_Sheet.xlsx'
wb.save(out_path)
print('saved', out_path)
